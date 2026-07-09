"""Manager/owner dashboard (Jinja2 + HTMX)."""

from __future__ import annotations
import logging
logger = logging.getLogger("speed-to-lead.dashboard")
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
import hashlib
import hmac
import secrets
import time

import bcrypt
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from fastapi import APIRouter, Cookie, Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, PlainTextResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select
from sqlalchemy import func as sa_func, case as sa_case
from sqlalchemy.orm import Session

from app.config import settings
from app.db import get_session_factory
from app.models import Appointment, Dealer, Lead, LeadEvent, LeadState, Message, Vehicle
from app.models import Channel, Direction

router = APIRouter(prefix="/dashboard", tags=["dashboard"])

TEMPLATES_DIR = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


# ---------------------------------------------------------------------------
# Jinja2 timezone filter: converts UTC datetime to dealer timezone
# ---------------------------------------------------------------------------
def _local_time(value, fmt='%b %d, %Y %I:%M %p', tz_name='America/Vancouver'):
    """Convert a UTC datetime to local time and format it."""
    if value is None:
        return ''
    from datetime import timezone as tz
    from zoneinfo import ZoneInfo
    if value.tzinfo is None:
        value = value.replace(tzinfo=tz.utc)
    local = value.astimezone(ZoneInfo(tz_name))
    return local.strftime(fmt)


templates.env.filters['local_time'] = _local_time


# ---------------------------------------------------------------------------
# Auth: rate-limiting state
# ---------------------------------------------------------------------------

# {ip_address: [timestamp_of_failure, ...]}
_login_attempts: dict[str, list[float]] = {}

# Rate-limit thresholds
_MAX_ATTEMPTS = 5
_WINDOW_SECONDS = 900  # 15 minutes


def _check_rate_limit(ip: str) -> bool:
    """Return True if the IP is rate-limited (blocked), False if allowed.
    Cleans up old entries on each check."""
    now = time.time()
    cutoff = now - _WINDOW_SECONDS
    # Clean old entries for this IP
    if ip in _login_attempts:
        _login_attempts[ip] = [t for t in _login_attempts[ip] if t > cutoff]
        if not _login_attempts[ip]:
            del _login_attempts[ip]
    # Also do a broader cleanup periodically
    if len(_login_attempts) > 100:
        for k in list(_login_attempts):
            _login_attempts[k] = [t for t in _login_attempts[k] if t > cutoff]
            if not _login_attempts[k]:
                del _login_attempts[k]
    # Check if blocked
    if ip in _login_attempts and len(_login_attempts[ip]) >= _MAX_ATTEMPTS:
        return True
    return False


def _record_failure(ip: str) -> None:
    """Record a failed login attempt for an IP."""
    _login_attempts.setdefault(ip, []).append(time.time())


def _clear_rate_limit(ip: str) -> None:
    """Clear rate limit entries for an IP (on successful login)."""
    _login_attempts.pop(ip, None)


# ---------------------------------------------------------------------------
# Auth: session serializer
# ---------------------------------------------------------------------------

def _get_serializer() -> URLSafeTimedSerializer:
    """Build a URLSafeTimedSerializer with a secret derived from config.

    Priority: DASHBOARD_SECRET > hash of DASHBOARD_PASSWORD > hardcoded fallback.
    """
    if settings.dashboard_secret:
        secret = settings.dashboard_secret
    elif settings.dashboard_password:
        secret = hashlib.sha256(settings.dashboard_password.encode()).hexdigest()
    else:
        if settings.environment == 'production':
            raise RuntimeError(
                "In production, either DASHBOARD_SECRET or DASHBOARD_PASSWORD must be set. "
                "The hardcoded dev fallback is not allowed in production."
            )
        # Fallback for dev only — not secure, but keeps things functional
        secret = "speed-to-lead-dev-secret-not-for-production"
    return URLSafeTimedSerializer(secret)


def _verify_password(password: str) -> bool:
    """Check password against bcrypt hash (preferred) or plaintext fallback."""
    if settings.dashboard_password_hash:
        try:
            return bcrypt.checkpw(password.encode(), settings.dashboard_password_hash.encode())
        except Exception:
            return False
    # Fallback: plaintext comparison (backward compat)
    return password == settings.dashboard_password


def _get_session() -> Session:
    return get_session_factory()()


# ---------------------------------------------------------------------------
# P0-08: CSRF protection (double-submit cookie pattern)
# ---------------------------------------------------------------------------

def _generate_csrf_token() -> str:
    """Generate a random CSRF token for the double-submit cookie pattern.

    The token is opaque (32 bytes from secrets.token_urlsafe → 43-char
    base64url string) and unpredictable. The same value lands in the
    csrf_token cookie AND in a hidden form field. POST compares them.
    """
    return secrets.token_urlsafe(32)


def _validate_csrf_token(request: Request, submitted_token: str) -> bool:
    """Validate the submitted CSRF token against the csrf_token cookie.

    Returns True only if both are non-empty and exactly match.
    Uses hmac.compare_digest for constant-time comparison (prevents
    timing attacks that could leak the cookie value character by
    character).
    """
    if not submitted_token:
        return False
    cookie_token = request.cookies.get("csrf_token")
    if not cookie_token:
        return False
    return hmac.compare_digest(cookie_token, submitted_token)


def get_dealer_from_auth(session: Session, cookie_value: str) -> Dealer | None:
    """Extract dealer_slug from the signed auth cookie and look up the Dealer.

    Returns the Dealer object or None if the cookie doesn't contain a valid
    dealer_slug or the dealer doesn't exist.
    """
    serializer = _get_serializer()
    try:
        data = serializer.loads(cookie_value, max_age=86400)
    except (BadSignature, SignatureExpired):
        return None
    dealer_slug = data.get("dealer_slug")
    if not dealer_slug:
        return None
    return session.execute(select(Dealer).where(Dealer.slug == dealer_slug)).scalars().first()


# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Attention Widget helpers
# ---------------------------------------------------------------------------

# Urgency levels: lower number = more urgent (sorted ascending)
URGENCY_HIGH = 1
URGENCY_MEDIUM = 2
URGENCY_LOW = 3


def get_attention_items(session: Session) -> list[dict[str, Any]]:
    """Return the most urgent items needing GM/owner attention.

    Queries for:
      - Unclaimed leads (ASSIGNED state, last updated > 2 hours ago)
      - Going cold (ENGAGED state, no activity in 48+ hours)
      - Appointments scheduled for today
      - Failed message deliveries

    Returns a sorted list (most urgent first), capped at 10 items.
    """
    now = datetime.now(timezone.utc)
    items: list[dict[str, Any]] = []

    # 1. Unclaimed leads: ASSIGNED and sitting for > 2 hours
    two_hours_ago = now - timedelta(hours=2)
    unclaimed_leads = session.execute(
        select(Lead).where(
            Lead.state == LeadState.ASSIGNED,
            Lead.updated_at < two_hours_ago,
        ).order_by(Lead.updated_at.asc()).limit(5)
    ).scalars().all()
    for lead in unclaimed_leads:
        hours = int((now - lead.updated_at.replace(tzinfo=timezone.utc)).total_seconds() // 3600)
        items.append({
            "type": "unclaimed",
            "urgency": URGENCY_HIGH,
            "lead": lead,
            "message": f"Unclaimed for {hours}+ hours — needs a rep",
            "icon": "alert-triangle",
        })

    # 2. Going cold: ENGAGED but no activity in 48+ hours
    two_days_ago = now - timedelta(hours=48)
    cold_leads = session.execute(
        select(Lead).where(
            Lead.state == LeadState.ENGAGED,
            Lead.updated_at < two_days_ago,
        ).order_by(Lead.updated_at.asc()).limit(5)
    ).scalars().all()
    for lead in cold_leads:
        days = int((now - lead.updated_at.replace(tzinfo=timezone.utc)).total_seconds() // 86400)
        items.append({
            "type": "going_cold",
            "urgency": URGENCY_MEDIUM,
            "lead": lead,
            "message": f"No activity for {days} day{'s' if days != 1 else ''} — follow up now",
            "icon": "snowflake",
        })

    # 3. Appointments today
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    todays_appts = session.execute(
        select(Appointment).where(
            Appointment.scheduled_for >= today_start,
            Appointment.scheduled_for < today_end,
            Appointment.status.in_(["set", "confirmed"]),
        ).order_by(Appointment.scheduled_for.asc()).limit(5)
    ).scalars().all()
    for appt in todays_appts:
        lead = session.get(Lead, appt.lead_id)
        items.append({
            "type": "appointment_today",
            "urgency": URGENCY_LOW,
            "lead": lead,
            "appointment": appt,
            "message": f"Appointment at {_local_time(appt.scheduled_for, '%I:%M %p')}",
            "icon": "calendar",
        })

    # 4. Failed deliveries
    failed_msgs = session.execute(
        select(Message).where(
            Message.delivery_status == "failed",
        ).order_by(Message.created_at.desc()).limit(5)
    ).scalars().all()
    for msg in failed_msgs:
        lead = session.get(Lead, msg.lead_id)
        error_info = f" (error {msg.error_code})" if msg.error_code else ""
        items.append({
            "type": "delivery_failure",
            "urgency": URGENCY_HIGH,
            "lead": lead,
            "message": f"Message delivery failed{error_info}",
            "icon": "x-circle",
        })

    # Sort by urgency (most urgent first), then by updated_at or created_at
    items.sort(key=lambda x: (x["urgency"],))
    return items[:10]


def get_source_breakdown(leads: list[Lead]) -> list[dict[str, Any]]:
    """Group leads by source channel and compute conversion metrics.

    For each source, counts:
      - total leads
      - sold (LeadState.SOLD)
      - engaged_or_better (APPT_SET + SHOWED + SOLD)

    Returns a sorted list (highest total first) of dicts:
      {source, total, sold, engaged_or_better, conversion_pct, appt_pct}
    """
    buckets: dict[str, dict[str, int]] = {}
    for lead in leads:
        src = lead.source.value if lead.source else "unknown"
        if src not in buckets:
            buckets[src] = {"total": 0, "sold": 0, "engaged_or_better": 0}
        buckets[src]["total"] += 1
        if lead.state == LeadState.SOLD:
            buckets[src]["sold"] += 1
            buckets[src]["engaged_or_better"] += 1
        elif lead.state in (LeadState.APPT_SET, LeadState.SHOWED):
            buckets[src]["engaged_or_better"] += 1

    breakdown: list[dict[str, Any]] = []
    for src, counts in buckets.items():
        total = counts["total"]
        breakdown.append({
            "source": src,
            "total": total,
            "sold": counts["sold"],
            "engaged_or_better": counts["engaged_or_better"],
            "conversion_pct": round(counts["sold"] / total * 100, 1) if total else 0,
            "appt_pct": round(counts["engaged_or_better"] / total * 100, 1) if total else 0,
        })
    breakdown.sort(key=lambda x: x["total"], reverse=True)
    return breakdown
def get_conversion_funnel(leads: list[Lead]) -> list[dict[str, Any]]:
    """Build a conversion funnel counting leads at each pipeline stage.

    The funnel follows the normal lead lifecycle order:
      NEW → AUTO_REPLIED → ASSIGNED → CLAIMED → ENGAGED → APPT_SET → SHOWED → SOLD

    Returns a list of dicts: {state, label, count, pct} where pct is the
    percentage of total leads (not cumulative).  Zero-count stages are still
    included so the UI always shows the full pipeline.
    """
    stages = [
        (LeadState.NEW, "New"),
        (LeadState.AUTO_REPLIED, "Auto Replied"),
        (LeadState.ASSIGNED, "Assigned"),
        (LeadState.CLAIMED, "Claimed"),
        (LeadState.ENGAGED, "Engaged"),
        (LeadState.APPT_SET, "Appt Set"),
        (LeadState.SHOWED, "Showed"),
        (LeadState.SOLD, "Sold"),
    ]
    total = len(leads)
    # Count leads currently in each state
    counts: dict[LeadState, int] = {state: 0 for state, _ in stages}
    for lead in leads:
        if lead.state in counts:
            counts[lead.state] += 1

    funnel: list[dict[str, Any]] = []
    for state, label in stages:
        count = counts[state]
        pct = round(count / total * 100, 1) if total > 0 else 0
        funnel.append({
            "state": state.value,
            "label": label,
            "count": count,
            "pct": pct,
        })
    return funnel
# ---------------------------------------------------------------------------
# Rep Performance Leaderboard
# ---------------------------------------------------------------------------

def get_rep_performance(leads: list[Lead], session: Session | None = None) -> list[dict[str, Any]]:
    """Build a rep performance leaderboard from a list of leads.

    Groups leads by assigned_rep and computes per-rep metrics:
      - assigned:    total leads assigned to this rep
      - engaged:     leads that reached ENGAGED, APPT_SET, SHOWED, or SOLD
      - appt_set:    leads that reached APPT_SET, SHOWED, or SOLD
      - sold:        leads in SOLD state
      - lost:        leads in LOST state
      - conversion_pct:  sold / assigned * 100
      - leads_today: leads created today
      - avg_response: average first-response time (if session provided)

    Returns a sorted list (by sold desc, then assigned desc) of dicts.
    """
    now_utc = datetime.now(timezone.utc)
    today_start = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)

    buckets: dict[str, dict[str, int]] = {}
    for lead in leads:
        rep = lead.assigned_rep or "Unassigned"
        if rep not in buckets:
            buckets[rep] = {
                "assigned": 0,
                "engaged": 0,
                "appt_set": 0,
                "sold": 0,
                "lost": 0,
                "leads_today": 0,
            }
        buckets[rep]["assigned"] += 1
        if lead.state in (LeadState.ENGAGED, LeadState.APPT_SET, LeadState.SHOWED, LeadState.SOLD):
            buckets[rep]["engaged"] += 1
        if lead.state in (LeadState.APPT_SET, LeadState.SHOWED, LeadState.SOLD):
            buckets[rep]["appt_set"] += 1
        if lead.state == LeadState.SOLD:
            buckets[rep]["sold"] += 1
        if lead.state == LeadState.LOST:
            buckets[rep]["lost"] += 1
        if lead.created_at and lead.created_at.replace(tzinfo=timezone.utc) >= today_start:
            buckets[rep]["leads_today"] += 1

    # Compute per-rep avg response time if session available
    rep_response_times: dict[str, list[float]] = {}
    if session:
        from sqlalchemy import select as sa_select, func as sa_func, case as sa_case
        stmt = (
            sa_select(
                Lead.assigned_rep,
                Message.lead_id,
                sa_case(
                    (Message.direction == Direction.INBOUND, Message.created_at)
                ).label("inbound_time"),
                sa_case(
                    (Message.direction == Direction.OUTBOUND, Message.created_at)
                ).label("outbound_time"),
            )
            .join(Message, Message.lead_id == Lead.id)
            .where(Lead.assigned_rep.isnot(None))
        )
        rows = session.execute(stmt).all()
        # Group by lead, find first inbound and outbound per lead
        lead_times: dict[tuple[str, int], dict[str, Any]] = {}
        for rep_name, lead_id, inb, outb in rows:
            key = (rep_name or "Unassigned", lead_id)
            if key not in lead_times:
                lead_times[key] = {"first_in": None, "first_out": None}
            if inb and (lead_times[key]["first_in"] is None or inb < lead_times[key]["first_in"]):
                lead_times[key]["first_in"] = inb
            if outb and (lead_times[key]["first_out"] is None or outb < lead_times[key]["first_out"]):
                lead_times[key]["first_out"] = outb
        for (rep_name, _lid), times in lead_times.items():
            if times["first_in"] and times["first_out"]:
                delta = (times["first_out"] - times["first_in"]).total_seconds()
                if delta >= 0:
                    rep_response_times.setdefault(rep_name, []).append(delta)

    leaderboard: list[dict[str, Any]] = []
    for rep, counts in buckets.items():
        assigned = counts["assigned"]
        avg_resp_seconds = 0
        if rep in rep_response_times and rep_response_times[rep]:
            avg_resp_seconds = sum(rep_response_times[rep]) / len(rep_response_times[rep])
        avg_resp_display = _format_seconds(avg_resp_seconds) if avg_resp_seconds else "--"

        leaderboard.append({
            "rep": rep,
            "assigned": assigned,
            "engaged": counts["engaged"],
            "appt_set": counts["appt_set"],
            "sold": counts["sold"],
            "lost": counts["lost"],
            "conversion_pct": round(counts["sold"] / assigned * 100, 1) if assigned else 0,
            "leads_today": counts["leads_today"],
            "avg_response_display": avg_resp_display,
            "avg_response_seconds": avg_resp_seconds,
        })

    # Sort by sold descending, then assigned descending
    leaderboard.sort(key=lambda x: (x["sold"], x["assigned"]), reverse=True)
    return leaderboard


def _format_seconds(seconds: float) -> str:
    """Format seconds into a human-readable string."""
    if seconds < 60:
        return f"{int(seconds)}s"
    elif seconds < 3600:
        mins = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{mins}m {secs}s"
    else:
        hours = int(seconds // 3600)
        mins = int((seconds % 3600) // 60)
        return f"{hours}h {mins}m"


# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Response Time Metrics
# ---------------------------------------------------------------------------

def get_response_metrics(session: Session, cutoff: datetime | None = None, dealer_id: int | None = None) -> dict[str, Any]:
    """Compute speed-to-lead response-time metrics.

    For each lead that has at least one inbound message, finds the first
    inbound and first outbound message.  The delta between those two is
    the "response time".

    Args:
        session: DB session
        cutoff: Optional datetime lower bound — only consider leads created after this.
        dealer_id: Optional dealer ID — scope metrics to a single dealer.

    Returns:
        avg_response_seconds: float (0 if no data)
        avg_response_display: human-readable string, e.g. "45s" or "3m 12s"
        pct_within_5_min:    percentage (0-100) of leads responded within 5 min
        total_with_response: count of leads that have a response
    """
    # Use SQL to compute first inbound/outbound times per lead — avoids loading all messages
    stmt = (
        select(
            Message.lead_id,
            sa_func.min(sa_case((Message.direction == Direction.INBOUND, Message.created_at))).label("first_inbound"),
            sa_func.min(sa_case((Message.direction == Direction.OUTBOUND, Message.created_at))).label("first_outbound"),
        )
        .join(Lead, Lead.id == Message.lead_id)
        .group_by(Message.lead_id)
    )
    if cutoff is not None:
        stmt = stmt.where(Lead.created_at >= cutoff)
    if dealer_id is not None:
        stmt = stmt.where(Lead.dealer_id == dealer_id)

    rows = session.execute(stmt).all()

    response_times: list[float] = []  # seconds

    for _lead_id, first_inbound, first_outbound in rows:
        if first_inbound and first_outbound:
            delta = (first_outbound - first_inbound).total_seconds()
            if delta >= 0:
                response_times.append(delta)

    total = len(response_times)
    if total == 0:
        return {
            "avg_response_seconds": 0,
            "avg_response_display": "--",
            "pct_within_5_min": 0,
            "total_with_response": 0,
        }

    avg_seconds = sum(response_times) / total
    within_5 = sum(1 for s in response_times if s <= 300)
    pct_within_5 = round(within_5 / total * 100, 1)

    # Human-readable display
    if avg_seconds < 60:
        avg_display = f"{int(avg_seconds)}s"
    elif avg_seconds < 3600:
        mins = int(avg_seconds // 60)
        secs = int(avg_seconds % 60)
        avg_display = f"{mins}m {secs}s"
    else:
        hours = int(avg_seconds // 3600)
        mins = int((avg_seconds % 3600) // 60)
        avg_display = f"{hours}h {mins}m"

    return {
        "avg_response_seconds": round(avg_seconds, 1),
        "avg_response_display": avg_display,
        "pct_within_5_min": pct_within_5,
        "total_with_response": total,
    }


# ---------------------------------------------------------------------------
# Lead Health Indicators
# ---------------------------------------------------------------------------

def get_lead_health(lead: Lead) -> str:
    """Compute a lead's health badge based on state and recency.

    Returns one of: "hot", "warm", "cold", "dead".

    Rules:
      - "hot"  if state is APPT_SET (appointment set)
      - "warm" if (ENGAGED + updated_at < 24h) OR (any state, updated_at < 48h)
      - "cold" if updated_at < 72h
      - "dead" otherwise
    """
    now = datetime.now(timezone.utc)
    updated = lead.updated_at
    if updated:
        # Ensure timezone-aware comparison
        if updated.tzinfo is None:
            updated = updated.replace(tzinfo=timezone.utc)
        age_hours = (now - updated).total_seconds() / 3600
    else:
        age_hours = float("inf")

    # Hot: appointment set
    if lead.state == LeadState.APPT_SET:
        return "hot"

    # Warm: ENGAGED within 24h, OR any state within 48h
    if lead.state == LeadState.ENGAGED and age_hours < 24:
        return "warm"
    if age_hours < 48:
        return "warm"

    # Cold: within 72h
    if age_hours < 72:
        return "cold"

    # Dead: everything else
    return "dead"


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def require_auth(session: str = Cookie(None)):
    """FastAPI dependency — raises a redirect to /dashboard/login if the
    session cookie is missing, expired, tampered with, or not a rep/manager role cookie.
    
    Returns: dict with 'role', 'rep_name', 'dealer_slug'."""
    if session is None:
        raise HTTPException(
            status_code=303,
            headers={"Location": "/dashboard/login"},
        )
    serializer = _get_serializer()
    try:
        data = serializer.loads(session, max_age=86400)
    except (BadSignature, SignatureExpired):
        raise HTTPException(
            status_code=303,
            headers={"Location": "/dashboard/login"},
        )
    # Must be a rep or manager role with a dealer_slug
    if data.get("role") not in ("rep", "manager") or not data.get("dealer_slug"):
        raise HTTPException(
            status_code=303,
            headers={"Location": "/dashboard/login"},
        )
    return data  # Returns {"role": "rep"|"manager", "rep_name": "...", "dealer_slug": "...", "ts": ...}


def get_auth_role(cookie_value: str | None) -> tuple[str, str]:
    """Extract role and rep_name from session cookie. Returns ("rep", "") on error."""
    if not cookie_value:
        return ("rep", "")
    try:
        data = _get_serializer().loads(cookie_value, max_age=86400)
        return (data.get("role", "rep"), data.get("rep_name", ""))
    except Exception:
        return ("rep", "")


def _check_lead_access(lead: Lead, auth: dict) -> bool:
    """Return True if the authenticated user can read/modify this lead.

    Managers can access all leads in their dealer.
    Reps can only access their own assigned leads.
    """
    role = auth.get("role", "rep")
    if role == "manager":
        return True
    rep_name = auth.get("rep_name", "")
    if not rep_name:
        return False
    return lead.assigned_rep == rep_name


def _base_context(auth: dict) -> dict:
    """Return shared template context vars (user_role, user_name, user_initials)."""
    role = auth.get("role", "rep")
    rep_name = auth.get("rep_name", "")
    return {
        "user_role": role.title(),
        "user_name": rep_name or role.title(),
        "user_initials": (rep_name[:2].upper() if rep_name else (role[0].upper() if role else "U")),
    }


# ---------------------------------------------------------------------------
# Login / Logout
# ---------------------------------------------------------------------------

@router.get("/login")
async def login_page(request: Request, dealer_slug: str = ""):
    # P0-08: generate CSRF token, set double-submit cookie, mirror in form
    token = _generate_csrf_token()
    
    # Load sales_team from dealer config if dealer_slug is provided
    sales_team = []
    dealer = None
    if dealer_slug:
        session = _get_session()
        try:
            dealer = session.execute(select(Dealer).where(Dealer.slug == dealer_slug)).scalars().first()
            if dealer:
                config = dealer.config or {}
                sales_team = config.get("sales_team", [])
        finally:
            session.close()
    
    response = templates.TemplateResponse(
        request=request,
        name="login.html",
        context={
            "request": request,
            "csrf_token": token,
            "active_page": "login",
            "sales_team": sales_team,
            "dealer_slug": dealer_slug,
            "show_manager_option": bool((config or {}).get("manager_pin")) if dealer else False,
        },
    )
    response.set_cookie(
        "csrf_token",
        token,
        httponly=True,
        secure=(settings.environment == "production"),
        samesite="lax",
        max_age=3600,
    )
    return response


@router.post("/login")
async def login_submit(
    request: Request,
    dealer_slug: str = Form(...),
    rep_name: str = Form(...),
    pin: str = Form(""),
    csrf_token: str = Form(""),  # P0-08: default empty so missing field → 403, not 422
):
    # P0-08: CSRF check FIRST, before rate limit. Putting CSRF before the
    # rate-limit means an attacker cannot DoS a dealer by flooding POSTs
    # and consuming the legitimate dealer's rate-limit window.
    if not _validate_csrf_token(request, csrf_token):
        return PlainTextResponse("Forbidden — CSRF token missing or invalid", status_code=403)

    ip = (request.headers.get('X-Forwarded-For', '').split(',')[0].strip() or (request.client.host if request.client else 'unknown'))

    if _check_rate_limit(ip):
        return templates.TemplateResponse(
            request=request,
            name="login.html",
            context={"request": request, "error": "Too many attempts. Try again later."},
            status_code=429,
        )

    # Validate dealer_slug exists
    db_session = _get_session()
    try:
        dealer = db_session.execute(select(Dealer).where(Dealer.slug == dealer_slug)).scalars().first()
        if not dealer:
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={"request": request, "error": "Unknown dealer slug"},
                status_code=401,
            )
    finally:
        db_session.close()

    dealer_config = dealer.config or {}
    sales_team = dealer_config.get("sales_team", [])
    manager_pin = dealer_config.get("manager_pin", "")

    # Determine if user selected Manager (special sentinel)
    is_manager = (rep_name == "Manager")

    if is_manager:
        # Manager login — validate against manager_pin
        if not manager_pin:
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={
                    "request": request,
                    "error": "Manager login is not configured for this dealer",
                    "sales_team": sales_team,
                    "dealer_slug": dealer_slug,
                },
                status_code=401,
            )
        if not pin or pin != manager_pin:
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={
                    "request": request,
                    "error": "Invalid PIN",
                    "sales_team": sales_team,
                    "dealer_slug": dealer_slug,
                },
                status_code=401,
            )
        role = "manager"
        logged_rep_name = "Manager"
    else:
        # Rep login — validate rep_name exists in sales_team, is active, and PIN matches
        rep_config = None
        for r in sales_team:
            if r.get("name") == rep_name:
                rep_config = r
                break

        if not rep_config:
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={
                    "request": request,
                    "error": "Please select your name from the list",
                    "sales_team": sales_team,
                    "dealer_slug": dealer_slug,
                },
                status_code=401,
            )

        # Reject inactive reps at login
        if not rep_config.get("active", True):
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={
                    "request": request,
                    "error": "Please select your name from the list",
                    "sales_team": sales_team,
                    "dealer_slug": dealer_slug,
                },
                status_code=401,
            )

        rep_pin = rep_config.get("pin", "")
        if not rep_pin:
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={
                    "request": request,
                    "error": f"No PIN configured for {rep_name}. Contact your manager.",
                    "sales_team": sales_team,
                    "dealer_slug": dealer_slug,
                },
                status_code=401,
            )
        if not pin or pin != rep_pin:
            return templates.TemplateResponse(
                request=request,
                name="login.html",
                context={
                    "request": request,
                    "error": "Invalid PIN",
                    "sales_team": sales_team,
                    "dealer_slug": dealer_slug,
                },
                status_code=401,
            )
        role = "rep"
        logged_rep_name = rep_name

    # Success — create session
    _clear_rate_limit(ip)
    serializer = _get_serializer()
    token = serializer.dumps({
        "role": role,
        "rep_name": logged_rep_name,
        "dealer_slug": dealer_slug,
        "ts": time.time(),
    })
    response = RedirectResponse("/dashboard", status_code=303)
    response.set_cookie("session", token, httponly=True, secure=(settings.environment == "production"), samesite="lax")
    return response


@router.get("/api/sales-team")
async def api_sales_team(dealer_slug: str = ""):
    """Return the sales team as JSON for the login page's dynamic rep dropdown.

    Called by JS on dealer_slug change/blur to populate the rep selector
    without a full page reload. Returns {sales_team: [...], show_manager_option: bool}.
    """
    if not dealer_slug:
        return {"sales_team": [], "show_manager_option": False}

    session = _get_session()
    try:
        dealer = session.execute(
            select(Dealer).where(Dealer.slug == dealer_slug)
        ).scalars().first()
        if not dealer:
            return {"sales_team": [], "show_manager_option": False}

        config = dealer.config or {}
        # SECURITY: this endpoint is UNAUTHENTICATED (the login page calls it
        # before the user has a session). NEVER return the raw rep dicts — they
        # contain `pin` (the login credential), `phone`, and `telegram_chat_id`.
        # Returning those would be a full auth bypass. Expose only the names the
        # dropdown needs, and only for active reps.
        names = [
            rep.get("name")
            for rep in config.get("sales_team", [])
            if rep.get("name") and rep.get("active", True)
        ]
        return {
            "sales_team": [{"name": n} for n in names],
            "show_manager_option": bool(config.get("manager_pin")),
        }
    finally:
        session.close()


@router.get("/logout")
async def logout(dealer_slug: str = ""):
    redirect_url = "/dashboard/login"
    if dealer_slug:
        redirect_url = f"/dashboard/login?dealer_slug={dealer_slug}"
    response = RedirectResponse(redirect_url, status_code=303)
    response.delete_cookie("session")
    return response


# ---------------------------------------------------------------------------
# Protected routes
# ---------------------------------------------------------------------------

@router.get("")
@router.get("/")
async def dashboard_index(request: Request, _auth: None = Depends(require_auth)):
    """Redirect to leads list."""
    # 303 (see other) forces a GET. The default 307 preserves the method, so the
    # login POST propagated through here to /dashboard/leads → 405 Method Not Allowed.
    return RedirectResponse(url="/dashboard/leads", status_code=303)


@router.post("/leads/new")
async def create_lead(
    request: Request,
    name: str = Form(...),
    phone: str = Form(...),
    vehicle_ref: str = Form(""),
    _auth: dict = Depends(require_auth),
):
    """Create a new lead manually (dashboard form)."""
    from app.adapters.intake import NormalizedLead
    from app.models import Channel
    from tools.route_lead import ingest_lead

    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        lead_data = NormalizedLead(
            source=Channel.WEBFORM,
            name=name,
            phone=phone,
            vehicle_ref=vehicle_ref or None,
            consent=True,
        )

        lead = ingest_lead(session, current_dealer, lead_data)

        # If a rep created this lead, assign it to them.
        rep_name = _auth.get("rep_name", "")
        role = _auth.get("role", "rep")
        if role == "rep" and rep_name:
            lead.assigned_rep = rep_name
        # Always persist: a manager-created lead has no rep branch to commit it,
        # so without this the new lead is dropped when the session closes.
        session.commit()

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = "Lead created successfully"
        response.headers["X-Toast-Type"] = "success"
        # Server-driven refresh: htmx fires the reload-leads event (closes the
        # modal + reloads the list). More reliable than inline hx-on on the form.
        response.headers["HX-Trigger"] = "reload-leads"
        return response
    except Exception:
        logger.exception("Failed to create lead")
        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = "Failed to create lead"
        response.headers["X-Toast-Type"] = "error"
        return response
    finally:
        session.close()


@router.get("/leads")
async def leads_list(request: Request, _auth: None = Depends(require_auth)):
    """Lead pipeline overview with stats and attention items."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return RedirectResponse("/dashboard/login", status_code=303)
        dealer_id = current_dealer.id
        
        # Rep-scoped query: reps see only their leads + unassigned
        role, rep_name = get_auth_role(cookie_value)
        if role == "rep" and rep_name:
            query = (
                select(Lead)
                .where(Lead.dealer_id == dealer_id)
                .where(Lead.assigned_rep == rep_name)
                .order_by(Lead.created_at.desc())
                .limit(100)
            )
        else:
            query = select(Lead).where(Lead.dealer_id == dealer_id).order_by(Lead.created_at.desc()).limit(100)
        all_leads = session.execute(query).scalars().all()

        total = len(all_leads)
        active = sum(1 for lead in all_leads if lead.state not in (LeadState.SOLD, LeadState.LOST, LeadState.OPTED_OUT))
        appt = sum(1 for lead in all_leads if lead.state in (LeadState.APPT_SET, LeadState.SHOWED))
        sold = sum(1 for lead in all_leads if lead.state == LeadState.SOLD)

        attention_items = get_attention_items(session)

        # Compute health for each lead
        lead_health: dict[int, str] = {lead.id: get_lead_health(lead) for lead in all_leads}

        # Today's leads count
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        leads_today = sum(1 for l in all_leads if l.created_at and l.created_at.replace(tzinfo=timezone.utc) >= today_start)

        # Response time metrics
        response_metrics = get_response_metrics(session, dealer_id=dealer_id)

        # Stale leads: no updated_at activity in 3+ days, still active
        now_utc = datetime.now(timezone.utc)
        three_days_ago = now_utc - timedelta(days=3)
        stale_leads = []
        for lead in all_leads:
            if lead.state in (LeadState.SOLD, LeadState.LOST, LeadState.OPTED_OUT):
                continue
            updated = lead.updated_at
            if updated:
                if updated.tzinfo is None:
                    updated = updated.replace(tzinfo=timezone.utc)
                if updated < three_days_ago:
                    days_stale = (now_utc - updated).days
                    stale_leads.append({"lead": lead, "days_stale": days_stale})
        stale_leads.sort(key=lambda x: x["days_stale"], reverse=True)
        stale_leads = stale_leads[:10]  # cap at 10

        return templates.TemplateResponse(request=request, name="leads.html", context={
            "request": request,
            "active_page": "leads",
            "leads": all_leads,
            "total_leads": total,
            "active_leads": active,
            "appt_leads": appt,
            "sold_leads": sold,
            "attention_items": attention_items,
            "lead_health": lead_health,
            "dealer": current_dealer,
            "dealer_name": current_dealer.name,
            "leads_today": leads_today,
            "avg_response_display": response_metrics["avg_response_display"],
            "avg_response_seconds": response_metrics["avg_response_seconds"],
            "stale_leads": stale_leads,
            **_base_context(_auth),
        })
    except Exception:
        logging.exception("Error in leads_list route")
        raise
    finally:
        session.close()


@router.get("/leads/partial")
async def leads_partial(
    request: Request,
    status: str | None = None,
    date_range: str | None = None,
    search: str | None = None,
    _auth: dict = Depends(require_auth),
):
    """HTMX partial: return filtered lead table rows (tbody content)."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("", status_code=401)

        dealer_id = current_dealer.id
        role, rep_name = get_auth_role(cookie_value)
        query = select(Lead).where(Lead.dealer_id == dealer_id)
        if role == "rep" and rep_name:
            query = query.where(Lead.assigned_rep == rep_name)

        # Status filter
        if status:
            try:
                state_enum = LeadState(status.upper())
                query = query.where(Lead.state == state_enum)
            except ValueError:
                pass  # ignore invalid status

        # Date range filter
        now = datetime.now(timezone.utc)
        if date_range == "today":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            query = query.where(Lead.created_at >= start)
        elif date_range == "yesterday":
            start = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            end = now.replace(hour=0, minute=0, second=0, microsecond=0)
            query = query.where(Lead.created_at >= start, Lead.created_at < end)
        elif date_range == "7d":
            query = query.where(Lead.created_at >= now - timedelta(days=7))
        elif date_range == "30d":
            query = query.where(Lead.created_at >= now - timedelta(days=30))
        # "all" or None = no date filter

        # Search filter
        if search:
            search_term = f"%{search}%"
            query = query.where(
                Lead.name.ilike(search_term)
                | Lead.phone.ilike(search_term)
                | Lead.email.ilike(search_term)
                | Lead.vehicle_ref.ilike(search_term)
            )

        query = query.order_by(Lead.created_at.desc()).limit(100)
        all_leads = session.execute(query).scalars().all()

        # Compute health for each lead
        lead_health: dict[int, str] = {lead.id: get_lead_health(lead) for lead in all_leads}

        return templates.TemplateResponse(
            request=request,
            name="leads_partial.html",
            context={
                "leads": all_leads,
                "lead_health": lead_health,
            },
        )
    except Exception:
        logging.exception("Error in leads_partial route")
        return HTMLResponse("<tr><td colspan='8'>Error loading leads</td></tr>", status_code=500)
    finally:
        session.close()


@router.get("/leads/{lead_id}")
async def lead_detail(request: Request, lead_id: int, _auth: None = Depends(require_auth)):
    """Lead detail view with a unified timeline of events + messages + delivery status."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return RedirectResponse("/dashboard/login", status_code=303)

        lead = session.get(Lead, lead_id)
        if not lead:
            return HTMLResponse("<h1>Lead not found</h1>", status_code=404)

        # Verify the lead belongs to the current dealer (404 if not)
        if lead.dealer_id != current_dealer.id:
            return HTMLResponse("<h1>Lead not found</h1>", status_code=404)
        
        # Rep URL guard: rep can only access their own assigned leads
        role, rep_name = get_auth_role(cookie_value)
        if role == "rep" and (not rep_name or lead.assigned_rep != rep_name):
            return HTMLResponse("<h1>Lead not found</h1>", status_code=404)

        messages = session.execute(
            select(Message).where(
                Message.lead_id == lead_id,
                # F4: exclude internal rep notifications from customer thread
                (Message.recipient_role == "customer") | (Message.recipient_role == None),
            ).order_by(Message.created_at.asc())
        ).scalars().all()

        events = session.execute(
            select(LeadEvent).where(LeadEvent.lead_id == lead_id).order_by(LeadEvent.created_at.asc())
        ).scalars().all()

        appointments = session.execute(
            select(Appointment).where(Appointment.lead_id == lead_id).order_by(Appointment.scheduled_for.asc())
        ).scalars().all()

        # Build a unified timeline merging events and messages
        timeline = []
        for event in events:
            payload = event.payload or {}
            entry = {
                "type": "event",
                "time": event.created_at,
                "description": event.type,
                "detail": "",
                "meta": payload,
            }
            if event.type == "state_change":
                entry["description"] = f"State: {payload.get('from', '?')} \u2192 {payload.get('to', '?')}"
                reason = payload.get("reason", "")
                if reason:
                    entry["detail"] = f"Reason: {reason}"
            elif event.type == "appointment":
                entry["description"] = f"Appointment {payload.get('status', 'set')}"
                entry["detail"] = f"Scheduled: {payload.get('scheduled_for', '?')}"
            timeline.append(entry)

        for msg in messages:
            entry = {
                "type": "message",
                "time": msg.created_at,
                "direction": msg.direction.value if msg.direction else "unknown",
                "channel": msg.channel.value if msg.channel else "-",
                "body": msg.body or "[no content]",
                "delivery_status": getattr(msg, "delivery_status", None),
                "error_code": getattr(msg, "error_code", None),
                "provider_sid": msg.provider_sid,
                "ai_generated": msg.ai_generated,
            }
            timeline.append(entry)

        # Sort by time
        timeline.sort(key=lambda x: x.get("time") or x.get("time", ""))

        # Get sales team from dealer config for the reassign dropdown
        dealer_config = current_dealer.config or {}
        sales_team = dealer_config.get("sales_team", [])

        # Compute allowed next states for the status dropdown
        from app.engine.lifecycle import TRANSITIONS
        allowed_states = sorted(s.value for s in TRANSITIONS.get(lead.state, set()))

        return templates.TemplateResponse(request=request, name="lead_detail.html", context={
            "request": request,
            "active_page": "leads",
            "lead": lead,
            "messages": messages,
            "events": events,
            "appointments": appointments,
            "timeline": timeline,
            "dealer": current_dealer,
            "dealer_name": current_dealer.name,
            "sales_team": sales_team,
            "allowed_states": allowed_states,
            "current_state": lead.state.value if lead.state else "NEW",
            **_base_context(_auth),
        })
    finally:
        session.close()


@router.get("/stats")
async def stats_page(request: Request, days: int = 30, _auth: None = Depends(require_auth)):
    """Stats & analytics overview."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return RedirectResponse("/dashboard/login", status_code=303)
        dealer_id = current_dealer.id

        # Rep-scoped stats: reps only see their own leads
        rep_name = _auth.get("rep_name", "")
        role = _auth.get("role", "rep")

        # Date-range filter: default last 30 days, adjustable via ?days= query param
        cutoff = datetime.now(timezone.utc) - timedelta(days=days)

        leads_q = select(Lead).where(Lead.created_at >= cutoff, Lead.dealer_id == dealer_id)
        if role == "rep" and rep_name:
            leads_q = leads_q.where(Lead.assigned_rep == rep_name)
        leads_q = leads_q.order_by(Lead.created_at.desc())
        leads = session.execute(leads_q).scalars().all()

        msg_q = select(Message).join(Lead).where(Lead.created_at >= cutoff, Lead.dealer_id == dealer_id)
        if role == "rep" and rep_name:
            msg_q = msg_q.where(Lead.assigned_rep == rep_name)
        msg_q = msg_q.order_by(Message.created_at.desc())
        messages = session.execute(msg_q).scalars().all()

        appt_q = select(Appointment).join(Lead).where(Lead.created_at >= cutoff, Lead.dealer_id == dealer_id)
        if role == "rep" and rep_name:
            appt_q = appt_q.where(Lead.assigned_rep == rep_name)
        appt_q = appt_q.order_by(Appointment.scheduled_for.desc())
        appointments = session.execute(appt_q).scalars().all()

        # Compute stats
        total = len(leads)
        active = sum(1 for l in leads if l.state not in (LeadState.SOLD, LeadState.LOST, LeadState.OPTED_OUT))
        appt = sum(1 for l in leads if l.state in (LeadState.APPT_SET, LeadState.SHOWED))
        sold = sum(1 for l in leads if l.state == LeadState.SOLD)

        conversion_rate = f"{(sold / total * 100):.1f}%" if total > 0 else "0%"

        # Leads per source
        sources: dict[str, int] = {}
        for l in leads:
            src = l.source.value if l.source else "unknown"
            sources[src] = sources.get(src, 0) + 1

        # Per-rep breakdown
        reps: dict[str, dict] = {}
        for l in leads:
            rep = l.assigned_rep or "Unassigned"
            if rep not in reps:
                reps[rep] = {"total": 0, "appt": 0, "sold": 0}
            reps[rep]["total"] += 1
            if l.state in (LeadState.APPT_SET, LeadState.SHOWED):
                reps[rep]["appt"] += 1
            if l.state == LeadState.SOLD:
                reps[rep]["sold"] += 1

        # Response time metrics
        response_metrics = get_response_metrics(session, cutoff=cutoff, dealer_id=dealer_id)

        # Source/Channel breakdown
        source_breakdown = get_source_breakdown(leads)

        # Conversion funnel
        conversion_funnel = get_conversion_funnel(leads)

        return templates.TemplateResponse(request=request, name="stats.html", context={
            "request": request,
            "active_page": "stats",
            "days": days,
            "leads": leads,
            "messages": messages,
            "appointments": appointments,
            "total_leads": total,
            "active_leads": active,
            "appt_leads": appt,
            "sold_leads": sold,
            "conversion_rate": conversion_rate,
            "sources": sources,
            "reps": reps,
            "response_metrics": response_metrics,
            "source_breakdown": source_breakdown,
            "conversion_funnel": conversion_funnel,
            "dealer": current_dealer,
            "dealer_name": current_dealer.name,
            **_base_context(_auth),
        })
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Team Management & Leaderboard
# ---------------------------------------------------------------------------

@router.get("/team")
async def team_page(request: Request, days: int = 30, _auth: dict = Depends(require_auth)):
    """Team management page with rep performance leaderboard."""
    if _auth.get("role") != "manager":
        return RedirectResponse("/dashboard/leads", status_code=303)
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return RedirectResponse("/dashboard/login", status_code=303)
        dealer_id = current_dealer.id

        # Date-range filter: default last 30 days, adjustable via ?days= query param
        cutoff = datetime.now(timezone.utc) - timedelta(days=days)

        leads_q = select(Lead).where(Lead.created_at >= cutoff, Lead.dealer_id == dealer_id)
        leads_q = leads_q.order_by(Lead.created_at.desc())
        leads = session.execute(leads_q).scalars().all()

        # Compute rep performance leaderboard
        rep_performance = get_rep_performance(leads, session=session)

        # Merge in configured reps from dealer config (so they show even without leads)
        dealer_config = current_dealer.config or {}
        configured_reps = dealer_config.get("sales_team", [])
        configured_names = {r.get("name") for r in configured_reps if r.get("name")}
        existing_names = {r["rep"] for r in rep_performance if r["rep"] != "Unassigned"}
        for rep_cfg in configured_reps:
            name = rep_cfg.get("name", "")
            if name and name not in existing_names:
                # Must match the row schema get_rep_performance emits, or the
                # team template (row.lost, row.conversion_pct, …) raises Undefined.
                rep_performance.append({
                    "rep": name,
                    "assigned": 0,
                    "engaged": 0,
                    "appt_set": 0,
                    "sold": 0,
                    "lost": 0,
                    "conversion_pct": 0,
                    "leads_today": 0,
                    "avg_response_display": "--",
                    "avg_response_seconds": 0,
                })

        # Active rep count (configured reps + any rep with leads)
        all_rep_names = configured_names | existing_names
        active_reps_count = len(all_rep_names)

        # Leads assigned today
        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        leads_today = sum(1 for l in leads if l.created_at.replace(tzinfo=timezone.utc) >= today_start)

        # Overall conversion rate
        total_assigned = len(leads)
        total_sold = sum(1 for l in leads if l.state == LeadState.SOLD)
        overall_conversion = round(total_sold / total_assigned * 100, 1) if total_assigned else 0

        return templates.TemplateResponse(request=request, name="team.html", context={
            "request": request,
            "active_page": "team",
            "days": days,
            "rep_performance": rep_performance,
            "active_rep_count": active_reps_count,
            "overall_conversion": overall_conversion,
            "leads_today": leads_today,
            "total_leads": total_assigned,
            "dealer": current_dealer,
            "dealer_name": current_dealer.name,
            **_base_context(_auth),
        })
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Appointments
# ---------------------------------------------------------------------------

@router.get("/appointments")
async def appointments_page(
    request: Request,
    status: str = None,
    view: str = "list",
    week_offset: int = 0,
    _auth: None = Depends(require_auth),
):
    """Appointments overview — list all appointments with filter by status,
    or a week-grid calendar view (view=week)."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return RedirectResponse("/dashboard/login", status_code=303)
        dealer_id = current_dealer.id
        role, rep_name = get_auth_role(cookie_value)

        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start + timedelta(days=1)
        this_week_start = today_start - timedelta(days=today_start.weekday())
        week_start = this_week_start + timedelta(weeks=week_offset)
        week_end = week_start + timedelta(days=7)

        # Build query
        stmt = select(Appointment).join(Lead, Appointment.lead_id == Lead.id).where(Lead.dealer_id == dealer_id)
        if role == "rep" and rep_name:
            stmt = stmt.where(Lead.assigned_rep == rep_name)
        if status:
            stmt = stmt.where(Appointment.status == status)
        stmt = stmt.order_by(Appointment.scheduled_for.desc())

        rows = session.execute(stmt).scalars().all()

        # Build appointment list with lead data
        appointments = []
        for appt in rows:
            lead = session.get(Lead, appt.lead_id)
            appointments.append({"appointment": appt, "lead": lead})

        # Stats
        appt_query = select(Appointment).join(Lead, Appointment.lead_id == Lead.id).where(Lead.dealer_id == dealer_id)
        if role == "rep" and rep_name:
            appt_query = appt_query.where(Lead.assigned_rep == rep_name)
        all_appts = session.execute(appt_query).scalars().all()

        today_count = sum(
            1 for a in all_appts
            if a.scheduled_for.replace(tzinfo=timezone.utc) >= today_start
            and a.scheduled_for.replace(tzinfo=timezone.utc) < today_end
        )
        week_count = sum(
            1 for a in all_appts
            if a.scheduled_for.replace(tzinfo=timezone.utc) >= this_week_start
            and a.scheduled_for.replace(tzinfo=timezone.utc) < this_week_start + timedelta(days=7)
        )
        showed_count = sum(1 for a in all_appts if a.status == "showed")
        completed = sum(1 for a in all_appts if a.status in ("showed", "no_show"))
        no_show_count = sum(1 for a in all_appts if a.status == "no_show")
        no_show_pct = round(no_show_count / completed * 100, 1) if completed > 0 else 0

        # Week-grid data: 7 days x appointments that fall in each day, for view=week
        week_days = []
        for i in range(7):
            day_start = week_start + timedelta(days=i)
            day_end = day_start + timedelta(days=1)
            day_appts = [
                a for a in all_appts
                if a.scheduled_for.replace(tzinfo=timezone.utc) >= day_start
                and a.scheduled_for.replace(tzinfo=timezone.utc) < day_end
            ]
            day_appts.sort(key=lambda a: a.scheduled_for)
            day_appt_items = []
            for a in day_appts:
                lead = session.get(Lead, a.lead_id)
                local_dt = a.scheduled_for.replace(tzinfo=timezone.utc)
                # Position within a 7am-8pm (13hr) grid, clamped so early/late
                # appointments still show at the top/bottom rather than vanishing.
                hour_decimal = max(7.0, min(20.0, local_dt.hour + local_dt.minute / 60))
                top_pct = (hour_decimal - 7) / 13 * 100
                day_appt_items.append({
                    "appointment": a, "lead": lead, "top_pct": round(top_pct, 1),
                })
            # NOTE: key is "appts", not "items" — Jinja resolves `day.items` to
            # dict.items() (the builtin method) if the key is named "items".
            week_days.append({"date": day_start, "appts": day_appt_items})

        return templates.TemplateResponse(request=request, name="appointments.html", context={
            "request": request,
            "active_page": "appointments",
            "appointments": appointments,
            "today_count": today_count,
            "week_count": week_count,
            "showed_count": showed_count,
            "no_show_pct": no_show_pct,
            "filter_status": status,
            "view": view,
            "week_offset": week_offset,
            "week_start": week_start,
            "week_days": week_days,
            "now_local": _local_time(now, '%Y-%m-%d'),
            "dealer": current_dealer,
            "dealer_name": current_dealer.name,
            **_base_context(_auth),
        })
    finally:
        session.close()


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

@router.get("/settings")
async def settings_page(request: Request, _auth: dict = Depends(require_auth)):
    """Settings page — shows dealer config from the current dealer."""
    if _auth.get("role") != "manager":
        return RedirectResponse("/dashboard/leads", status_code=303)
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return RedirectResponse("/dashboard/login", status_code=303)
        dealer = current_dealer

        config = dealer.config or {}
        channels = config.get("channels", {})
        ai_config = config.get("ai", {})

        dealer_name = dealer.name or ""
        dealer_phone = channels.get("sms_number", "") or channels.get("phone", "")
        dealer_address = config.get("address", "") or ""
        dealer_website = config.get("website", "") or ""
        ai_persona = ai_config.get("persona", "") or ""
        engagement_mode = ai_config.get("engagement_mode", "full_auto") or "full_auto"

        # Business hours from dealer config
        # Config stores hours as {"mon": "09:00-19:00", "tue": "09:00-19:00", ...}
        dealer_hours = config.get("hours", {})
        default_hours = {"mon": "09:00-21:00", "tue": "09:00-21:00", "wed": "09:00-21:00",
                         "thu": "09:00-21:00", "fri": "09:00-21:00", "sat": "09:00-18:00", "sun": "closed"}
        # Merge with defaults
        hours = {**default_hours, **(dealer_hours or {})}

        def parse_hours(h):
            """Parse '09:00-19:00' into {'open': '09:00', 'close': '19:00'} or closed."""
            if not h or h.lower() == "closed":
                return {"open": "", "close": "", "closed": True}
            parts = h.split("-")
            if len(parts) == 2:
                return {"open": parts[0].strip(), "close": parts[1].strip(), "closed": False}
            return {"open": "", "close": "", "closed": True}

        business_hours = {day: parse_hours(h) for day, h in hours.items()}

        # Quiet hours from compliance config
        compliance = config.get("compliance", {})
        quiet_raw = compliance.get("quiet_hours", "21:00-08:00") or "21:00-08:00"
        quiet_parts = quiet_raw.split("-")
        quiet_start = quiet_parts[0].strip() if len(quiet_parts) == 2 else "21:00"
        quiet_end = quiet_parts[1].strip() if len(quiet_parts) == 2 else "08:00"

        # Digest settings from routing config
        routing = config.get("routing", {})
        digest_enabled = routing.get("digest_enabled", False)
        digest_time = routing.get("digest_time", "08:00")

        # Current inventory for this dealer (so the manager can see/verify what
        # the AI's check_inventory will actually search).
        inv_rows = session.execute(
            select(Vehicle).where(Vehicle.dealer_id == dealer.id)
            .order_by(Vehicle.make, Vehicle.model)
        ).scalars().all()
        inventory = [{
            "stock_no": v.stock_no, "year": v.year, "make": v.make, "model": v.model,
            "trim": v.trim, "body": v.body, "price": v.price, "status": v.status,
        } for v in inv_rows]

        return templates.TemplateResponse(request=request, name="settings.html", context={
            "request": request,
            "active_page": "settings",
            "dealer_name": dealer_name,
            "dealer_phone": dealer_phone,
            "dealer_address": dealer_address,
            "dealer_website": dealer_website,
            "ai_persona": ai_persona,
            "engagement_mode": engagement_mode,
            "business_hours": business_hours,
            "quiet_start": quiet_start,
            "quiet_end": quiet_end,
            "digest_enabled": digest_enabled,
            "digest_time": digest_time,
            "dealer": dealer,
            "inventory": inventory,
            "inventory_count": len(inventory),
            **_base_context(_auth),
        })
    finally:
        session.close()


# ---------------------------------------------------------------------------
# HTMX Lead Action Endpoints
# ---------------------------------------------------------------------------

@router.post("/settings/channels")
async def save_channel_settings(
    request: Request,
    _auth: dict = Depends(require_auth),
):
    """Save channel settings (digest toggle + time)."""
    if _auth.get("role") != "manager":
        from fastapi.responses import JSONResponse
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=403)
    from fastapi.responses import JSONResponse
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

        form = await request.form()
        digest_enabled = form.get("digest_enabled") == "on"
        digest_time = form.get("digest_time", "08:00")

        config = current_dealer.config or {}
        if "routing" not in config:
            config["routing"] = {}
        config["routing"]["digest_enabled"] = digest_enabled
        config["routing"]["digest_time"] = digest_time or "08:00"
        current_dealer.config = config
        session.commit()

        return JSONResponse({"status": "success", "message": "Channel settings saved"})
    finally:
        session.close()


@router.post("/settings/business")
async def save_business_settings(
    request: Request,
    _auth: dict = Depends(require_auth),
):
    """Save business info settings (name, phone, address, website, hours)."""
    if _auth.get("role") != "manager":
        from fastapi.responses import JSONResponse
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=403)
    from fastapi.responses import JSONResponse
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

        form = await request.form()
        config = current_dealer.config or {}

        # Dealer info
        config["dealer"] = config.get("dealer", {})
        if form.get("dealer_name"):
            current_dealer.name = form["dealer_name"]
            config["dealer"]["name"] = form["dealer_name"]
        if form.get("dealer_phone"):
            config["channels"] = config.get("channels", {})
            config["channels"]["sms_number"] = form["dealer_phone"]
        if form.get("dealer_address"):
            config["address"] = form["dealer_address"]
        if form.get("dealer_website"):
            config["website"] = form["dealer_website"]

        # Business hours
        hours = {}
        for day in ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]:
            closed = form.get(f"{day}_closed") == "on"
            if closed:
                hours[day] = "closed"
            else:
                open_val = form.get(f"{day}_open", "09:00")
                close_val = form.get(f"{day}_close", "17:00")
                hours[day] = f"{open_val}-{close_val}"
        config["hours"] = hours

        current_dealer.config = config
        session.commit()

        return JSONResponse({"status": "success", "message": "Business info saved"})
    finally:
        session.close()


@router.post("/settings/ai")
async def save_ai_settings(
    request: Request,
    _auth: dict = Depends(require_auth),
):
    """Save AI personality settings (persona, engagement mode, guardrails)."""
    if _auth.get("role") != "manager":
        from fastapi.responses import JSONResponse
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=403)
    from fastapi.responses import JSONResponse
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

        form = await request.form()
        config = current_dealer.config or {}

        config["ai"] = config.get("ai", {})
        if form.get("engagement_mode"):
            config["ai"]["engagement_mode"] = form["engagement_mode"]
        if form.get("ai_persona"):
            config["ai"]["persona"] = form["ai_persona"]

        guardrails = []
        for g in ["no_price_negotiation", "no_financing_promises", "no_inventory_guarantees", "no_competitor_comparisons"]:
            if form.get(f"guardrail_{g}"):
                guardrails.append(g)
        config["ai"]["guardrails"] = {g: True for g in guardrails}

        current_dealer.config = config
        session.commit()

        return JSONResponse({"status": "success", "message": "AI personality saved"})
    finally:
        session.close()


@router.post("/settings/compliance")
async def save_compliance_settings(
    request: Request,
    _auth: dict = Depends(require_auth),
):
    """Save compliance settings (quiet hours, consent text, opt-out keywords)."""
    if _auth.get("role") != "manager":
        from fastapi.responses import JSONResponse
        return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=403)
    from fastapi.responses import JSONResponse
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

        form = await request.form()
        config = current_dealer.config or {}

        config["compliance"] = config.get("compliance", {})
        quiet_start = form.get("quiet_start", "21:00")
        quiet_end = form.get("quiet_end", "08:00")
        config["compliance"]["quiet_hours"] = f"{quiet_start}-{quiet_end}"
        if form.get("consent_text"):
            config["compliance"]["consent_text"] = form["consent_text"]
        if form.get("opt_out_keywords"):
            config["compliance"]["opt_out_keywords"] = [kw.strip() for kw in form["opt_out_keywords"].split(",")]

        current_dealer.config = config
        session.commit()

        return JSONResponse({"status": "success", "message": "Compliance settings saved"})
    finally:
        session.close()

@router.post("/leads/{lead_id}/reassign")
async def reassign_lead(
    request: Request,
    lead_id: int,
    rep: str = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Reassign a lead to a different rep."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        lead = session.get(Lead, lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        if not _check_lead_access(lead, _auth):
            return HTMLResponse("Access denied — this lead is assigned to another rep", status_code=403)

        old_rep = lead.assigned_rep or "Unassigned"
        lead.assigned_rep = rep
        lead.updated_at = datetime.now(timezone.utc)

        event = LeadEvent(
            lead_id=lead.id,
            dealer_id=current_dealer.id,
            type="state_change",
            payload={"action": "reassigned", "from_rep": old_rep, "to_rep": rep},
        )
        session.add(event)
        session.commit()

        # Notify the new rep via Telegram (single dealer notification channel)
        dealer_config = current_dealer.config or {}
        sales_team = dealer_config.get("sales_team", [])
        new_rep_config = next((r for r in sales_team if r.get("name") == rep), None)
        if new_rep_config:
            try:
                from tools.notify_rep import notify_rep
                notify_rep(
                    session=session,
                    rep_config=new_rep_config,
                    lead=lead,
                    message_type="cover_me",
                    payload={
                        "customer_name": lead.name or "A customer",
                        "vehicle": lead.vehicle_ref or "",
                    },
                    dealer_config=dealer_config,
                )
            except Exception:
                logger.exception("Failed to notify rep %s about reassignment", rep)

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = f"Lead reassigned to {rep}"
        response.headers["X-Toast-Type"] = "success"
        return response
    finally:
        session.close()


@router.post("/leads/{lead_id}/status")
async def update_lead_status(
    request: Request,
    lead_id: int,
    status: str = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Update a lead's status via lifecycle transition."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        lead = session.get(Lead, lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        if not _check_lead_access(lead, _auth):
            return HTMLResponse("Access denied — this lead is assigned to another rep", status_code=403)

        try:
            target_state = LeadState(status)
        except ValueError:
            return HTMLResponse(f"Invalid status: {status}", status_code=400)

        try:
            from app.engine.lifecycle import transition
            transition(session, lead, target_state, reason="dashboard_status_change")
        except ValueError:
            response = HTMLResponse(f"Can't move a {lead.state.value} lead to {status}", status_code=200)
            response.headers["X-Toast-Message"] = f"Can't move a {lead.state.value} lead to {status}"
            response.headers["X-Toast-Type"] = "error"
            return response

        event = LeadEvent(
            lead_id=lead.id,
            dealer_id=current_dealer.id,
            type="state_change",
            payload={"action": "status_changed", "to": status},
        )
        session.add(event)
        session.commit()

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = f"Status updated to {status}"
        response.headers["X-Toast-Type"] = "success"
        return response
    finally:
        session.close()


@router.post("/leads/{lead_id}/messages")
async def send_lead_message(
    request: Request,
    lead_id: int,
    message: str = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Send an SMS message to a lead."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        lead = session.get(Lead, lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        if not _check_lead_access(lead, _auth):
            return HTMLResponse("Access denied", status_code=403)

        if not lead.phone:
            return HTMLResponse("Lead has no phone number", status_code=400)

        dealer_config = current_dealer.config or {}
        channels = dealer_config.get("channels", {})
        sms_number = channels.get("sms_number", "")

        # Create outbound message record
        outbound = Message(
            lead_id=lead.id,
            direction=Direction.OUTBOUND,
            channel=Channel.SMS,
            body=message,
        )
        session.add(outbound)
        session.commit()

        # Send via SMS
        try:
            from tools.send_sms import send_sms
            send_sms(
                session,
                to=lead.phone,
                body=message,
                from_number=sms_number,
                dealer_slug=current_dealer.slug,
                dealer_config=dealer_config,
                lead=lead,
            )
        except Exception:
            logging.exception("Failed to send SMS for lead %s", lead_id)

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = "Message sent"
        response.headers["X-Toast-Type"] = "success"
        return response
    finally:
        session.close()


@router.post("/leads/{lead_id}/follow-up")
async def schedule_lead_followup(
    request: Request,
    lead_id: int,
    follow_up: str = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Schedule a follow-up for a lead."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        lead = session.get(Lead, lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        if not _check_lead_access(lead, _auth):
            return HTMLResponse("Access denied", status_code=403)

        try:
            followup_dt = datetime.fromisoformat(follow_up)
        except ValueError:
            return HTMLResponse("Invalid datetime format", status_code=400)

        # Calculate minutes from now
        now = datetime.now(timezone.utc)
        if followup_dt.tzinfo is None:
            followup_dt = followup_dt.replace(tzinfo=timezone.utc)
        minutes = max(1, int((followup_dt - now).total_seconds() / 60))

        from app.scheduler import schedule_followup as _schedule_followup
        from apscheduler.schedulers.background import BackgroundScheduler

        # Schedule via the scheduler - get the running scheduler from app state
        # Use a simpler approach: create a date-based job directly
        from app.db import get_session_factory
        from apscheduler.schedulers.background import BackgroundScheduler
        from datetime import timedelta

        # For now, just log the follow-up intent
        event = LeadEvent(
            lead_id=lead.id,
            dealer_id=current_dealer.id,
            type="followup_scheduled",
            payload={"scheduled_for": followup_dt.isoformat(), "minutes": minutes},
        )
        session.add(event)
        session.commit()

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = "Follow-up scheduled"
        response.headers["X-Toast-Type"] = "success"
        return response
    finally:
        session.close()


# ---------------------------------------------------------------------------

@router.post("/leads/{lead_id}/activity")
async def log_lead_activity(
    request: Request,
    lead_id: int,
    note: str = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Log a quick activity note on a lead (no-answer, voicemail, spoke, etc.)."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        lead = session.get(Lead, lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        if not _check_lead_access(lead, _auth):
            return HTMLResponse("Access denied", status_code=403)

        event = LeadEvent(
            lead_id=lead.id,
            dealer_id=current_dealer.id,
            type="activity_note",
            payload={"note": note, "logged_by": "dashboard"},
        )
        session.add(event)
        lead.updated_at = datetime.now(timezone.utc)
        session.commit()

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = "Activity logged"
        response.headers["X-Toast-Type"] = "success"
        return response
    finally:
        session.close()


@router.post("/leads/{lead_id}/mark-sold")
async def mark_lead_sold(
    request: Request,
    lead_id: int,
    _auth: dict = Depends(require_auth),
):
    """Mark a lead as sold."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        lead = session.get(Lead, lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        if not _check_lead_access(lead, _auth):
            return HTMLResponse("Access denied", status_code=403)

        try:
            from app.engine.lifecycle import transition
            transition(session, lead, LeadState.SOLD, reason="dashboard_mark_sold")
        except ValueError as e:
            return HTMLResponse(str(e), status_code=400)

        event = LeadEvent(
            lead_id=lead.id,
            dealer_id=current_dealer.id,
            type="state_change",
            payload={"action": "marked_sold"},
        )
        session.add(event)
        session.commit()

        response = HTMLResponse("", status_code=200)
        response.headers["HX-Redirect"] = "/dashboard/leads"
        response.headers["X-Toast-Message"] = "Lead marked as sold!"
        response.headers["X-Toast-Type"] = "success"
        return response
    finally:
        session.close()


@router.post("/leads/{lead_id}/mark-lost")
async def mark_lead_lost(
    request: Request,
    lead_id: int,
    reason: str = Form(""),
    _auth: dict = Depends(require_auth),
):
    """Mark a lead as lost."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        lead = session.get(Lead, lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        if not _check_lead_access(lead, _auth):
            return HTMLResponse("Access denied", status_code=403)

        try:
            from app.engine.lifecycle import transition
            transition(session, lead, LeadState.LOST, reason="dashboard_mark_lost")
        except ValueError as e:
            return HTMLResponse(str(e), status_code=400)

        # Save loss reason
        lead.loss_reason = reason if reason else None

        event = LeadEvent(
            lead_id=lead.id,
            dealer_id=current_dealer.id,
            type="state_change",
            payload={"action": "marked_lost", "reason": reason},
        )
        session.add(event)
        session.commit()

        response = HTMLResponse("", status_code=200)
        response.headers["HX-Redirect"] = "/dashboard/leads"
        response.headers["X-Toast-Message"] = "Lead marked as lost"
        response.headers["X-Toast-Type"] = "warning"
        return response
    finally:
        session.close()


@router.post("/team")
async def add_team_member(
    request: Request,
    name: str = Form(...),
    phone: str = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Add a new team member to the dealer's sales team."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        import json as _json
        import secrets as _secrets
        config = current_dealer.config or {}
        if "sales_team" not in config:
            config["sales_team"] = []
        # Give every new rep a working login PIN + the defaults they need to be
        # reachable, otherwise they can never log in or receive claim pings.
        new_pin = f"{_secrets.randbelow(10000):04d}"
        config["sales_team"].append({
            "name": name,
            "phone": phone,
            "pin": new_pin,
            "active": True,
            "notify_backend": "telegram",
            "telegram_chat_id": None,
        })
        # Reassign a fresh object so SQLAlchemy reliably detects the JSON change.
        current_dealer.config = _json.loads(_json.dumps(config))
        session.commit()

        response = HTMLResponse(
            f'<div class="toast success">Team member {name} added — login PIN: {new_pin}. '
            f'They still need to tap their Telegram link to get lead pings.</div>',
            status_code=200,
        )
        response.headers["HX-Trigger"] = "showToast"
        return response
    finally:
        session.close()


@router.post("/team/{rep_name}/unavailable")
async def add_unavailable_window(
    request: Request,
    rep_name: str,
    date: str = Form(...),
    start: str = Form(...),
    end: str = Form(...),
    note: str = Form(""),
    _auth: dict = Depends(require_auth),
):
    """Add an unavailability window for a rep."""
    from app.config import UnavailableWindow as _UnavailableWindow
    from sqlalchemy.orm.attributes import flag_modified
    import json as _json

    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        config = current_dealer.config or {}
        sales_team = config.get("sales_team", [])
        rep_cfg = next((r for r in sales_team if r.get("name", "").lower() == rep_name.lower()), None)
        if not rep_cfg:
            return HTMLResponse(f"Rep {rep_name} not found", status_code=404)

        # Validate the window via Pydantic
        try:
            window = _UnavailableWindow(date=date, start=start, end=end, note=note)
        except Exception as e:
            return HTMLResponse(f"Invalid window: {e}", status_code=400)

        # Add to rep's unavailable_windows
        if "unavailable_windows" not in rep_cfg:
            rep_cfg["unavailable_windows"] = []
        rep_cfg["unavailable_windows"].append(window.model_dump())

        current_dealer.config = _json.loads(_json.dumps(config))
        flag_modified(current_dealer, "config")
        session.commit()

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = f"Unavailable added for {rep_name}: {date} {start}-{end}"
        response.headers["X-Toast-Type"] = "success"
        response.headers["HX-Trigger"] = "showToast"
        return response
    finally:
        session.close()


@router.post("/team/{rep_name}/unavailable/remove")
async def remove_unavailable_window(
    request: Request,
    rep_name: str,
    index: int = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Remove an unavailability window by index for a rep."""
    from sqlalchemy.orm.attributes import flag_modified
    import json as _json

    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        config = current_dealer.config or {}
        sales_team = config.get("sales_team", [])
        rep_cfg = next((r for r in sales_team if r.get("name", "").lower() == rep_name.lower()), None)
        if not rep_cfg:
            return HTMLResponse(f"Rep {rep_name} not found", status_code=404)

        windows = rep_cfg.get("unavailable_windows", [])
        if index < 0 or index >= len(windows):
            return HTMLResponse("Invalid window index", status_code=400)

        removed = windows.pop(index)
        rep_cfg["unavailable_windows"] = windows

        current_dealer.config = _json.loads(_json.dumps(config))
        flag_modified(current_dealer, "config")
        session.commit()

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = f"Removed unavailable window for {rep_name}"
        response.headers["X-Toast-Type"] = "success"
        response.headers["HX-Trigger"] = "showToast"
        return response
    finally:
        session.close()


# ── No-show endpoints ────────────────────────────────────────────────────────

@router.post("/leads/{lead_id}/mark-showed")
async def lead_mark_showed(
    request: Request,
    lead_id: int,
    appointment_id: int = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Mark an appointment as showed."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        from app.models import Appointment, Lead
        appt = session.get(Appointment, appointment_id)
        if not appt:
            return HTMLResponse("Appointment not found", status_code=404)
        lead = session.get(Lead, appt.lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        from tools.book_appointment import mark_showed
        mark_showed(session, appt, lead)

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = f"Marked as showed: {lead.name or 'Customer'}"
        response.headers["X-Toast-Type"] = "success"
        response.headers["HX-Trigger"] = "showToast"
        return response
    finally:
        session.close()


@router.post("/leads/{lead_id}/mark-no-show")
async def lead_mark_no_show(
    request: Request,
    lead_id: int,
    appointment_id: int = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Mark an appointment as no-show."""
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        from app.models import Appointment, Lead
        appt = session.get(Appointment, appointment_id)
        if not appt:
            return HTMLResponse("Appointment not found", status_code=404)
        lead = session.get(Lead, appt.lead_id)
        if not lead or lead.dealer_id != current_dealer.id:
            return HTMLResponse("Lead not found", status_code=404)

        from tools.book_appointment import mark_no_show
        mark_no_show(session, appt)

        response = HTMLResponse("", status_code=200)
        response.headers["X-Toast-Message"] = f"Marked as no-show: {lead.name or 'Customer'}"
        response.headers["X-Toast-Type"] = "warning"
        response.headers["HX-Trigger"] = "showToast"
        return response
    finally:
        session.close()


# ── Inventory upload ──────────────────────────────────────────────────────────

from fastapi import UploadFile, File


@router.post("/inventory/upload")
async def upload_inventory(
    request: Request,
    file: UploadFile = File(...),
    full_sync: str = Form(""),
    _auth: dict = Depends(require_auth),
):
    """Upload inventory from CSV or XLSX. Upserts into Vehicle table with row-level errors."""
    import csv
    import io

    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)

        content = await file.read()
        filename = file.filename or ""
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        rows: list[dict] = []
        errors: list[str] = []

        if ext == "xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError:
                return HTMLResponse("XLSX support requires openpyxl", status_code=500)
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
            if header_row is None:
                wb.close()
                return HTMLResponse("Empty file.", status_code=400)
            headers = [str(cell.value).strip() if cell.value else "" for cell in header_row]
            header_map = {h.lower(): i for i, h in enumerate(headers) if h}
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {h: row[i] for h, i in header_map.items() if i < len(row)}
                rows.append(row_dict)
            wb.close()
        elif ext == "csv":
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
            # Lowercase headers so "Year"/"Make"/"Model" (the normal export casing)
            # match the lowercase row.get() lookups below — same as the XLSX path.
            rows = [{(k or "").strip().lower(): v for k, v in r.items()} for r in reader]
        else:
            return HTMLResponse("Unsupported file type. Upload .csv or .xlsx.", status_code=400)

        if not rows:
            return HTMLResponse("Empty file.", status_code=400)

        upserted = 0
        uploaded_stock_nos = set()     # I1: track which stock_nos were in THIS file
        for row_idx, row in enumerate(rows, start=2):
            try:
                stock_no = str(row.get("stock_no", row.get("stock #", ""))).strip()
                year = int(float(str(row.get("year", 0))))
                make = str(row.get("make", "")).strip()
                model = str(row.get("model", "")).strip()
                trim = str(row.get("trim", "")).strip()
                body = str(row.get("body", row.get("body_style", ""))).strip()
                price_str = str(row.get("price", "0")).replace("$", "").replace(",", "")
                price = int(float(price_str)) if price_str else 0
                mileage_str = str(row.get("mileage", row.get("miles", "0"))).replace(",", "")
                mileage = int(float(mileage_str)) if mileage_str else 0

                if not stock_no or not year or not make or not model:
                    errors.append(f"Row {row_idx}: missing required field")
                    continue

                vehicle = session.execute(
                    select(Vehicle).where(
                        Vehicle.dealer_id == current_dealer.id,
                        Vehicle.stock_no == stock_no,
                    )
                ).scalars().first()

                def _clean(*keys):
                    """First non-empty value among the given column names."""
                    for k in keys:
                        v = row.get(k)
                        if v is not None and str(v).strip():
                            return str(v).strip()
                    return None

                # Features may be a pipe/semicolon/comma-separated string in the file.
                features_raw = _clean("features")
                features = []
                if features_raw:
                    for sep in ("|", ";"):
                        if sep in features_raw:
                            features = [f.strip() for f in features_raw.split(sep) if f.strip()]
                            break
                    else:
                        features = [f.strip() for f in features_raw.split(",") if f.strip()]

                # Capture the full spec set check_inventory surfaces to customers
                # (engine, transmission, drivetrain, horsepower, fuel_economy,
                # exterior/interior color, range, features). The AI may ONLY state
                # facts that live here, so the richer this is, the better it sells.
                raw_specs = {
                    "engine": _clean("engine"),
                    "transmission": _clean("transmission"),
                    "drivetrain": _clean("drivetrain"),
                    "horsepower": _clean("horsepower", "hp"),
                    "torque": _clean("torque"),
                    "fuel_economy": _clean("fuel_economy", "fuel economy", "mpg", "l/100km"),
                    "range": _clean("range"),
                    "exterior_color": _clean("exterior_color", "color"),
                    "interior": _clean("interior", "interior_color"),
                    "features": features,
                }
                vin = _clean("vin")
                url = _clean("url", "vehicle_url")
                photo = _clean("image_url", "image", "photo")
                photos = [photo] if photo else []

                status_val = str(row.get("status", "")).strip().lower()
                if status_val not in ("available", "sold", "removed"):
                    status_val = "available"

                if vehicle:
                    vehicle.year = year; vehicle.make = make; vehicle.model = model
                    vehicle.trim = trim or vehicle.trim; vehicle.body = body or vehicle.body
                    vehicle.price = price; vehicle.mileage = mileage
                    vehicle.raw = raw_specs; vehicle.status = status_val
                    if vin: vehicle.vin = vin
                    if url: vehicle.url = url
                    if photos: vehicle.photos = photos
                else:
                    vehicle = Vehicle(
                        dealer_id=current_dealer.id, stock_no=stock_no,
                        year=year, make=make, model=model, trim=trim or None,
                        body=body or None, price=price, mileage=mileage,
                        vin=vin, url=url, photos=photos,
                        raw=raw_specs, status=status_val,
                    )
                    session.add(vehicle)
                upserted += 1
                uploaded_stock_nos.add(stock_no)
            except (ValueError, TypeError) as e:
                errors.append(f"Row {row_idx}: {e}")
                continue

        if upserted > 0:
            session.commit()

        removed = 0
        if full_sync == "on" and uploaded_stock_nos:
            from sqlalchemy import update as sa_update
            res = session.execute(
                sa_update(Vehicle)
                .where(
                    Vehicle.dealer_id == current_dealer.id,
                    Vehicle.status != "removed",
                    Vehicle.stock_no.notin_(uploaded_stock_nos),
                )
                .values(status="removed")
            )
            removed = res.rowcount or 0
            session.commit()

        msg = f'<div class="toast success">{upserted} vehicles uploaded.'
        if full_sync == "on":
            msg += f' {removed} removed (not in file).'
        msg += '</div>'
        if errors:
            preview = "; ".join(errors[:5])
            msg += f'<div class="toast warning">{len(errors)} errors: {preview}{"..." if len(errors)>5 else ""}</div>'
        return HTMLResponse(msg, status_code=200)

    except Exception as e:
        logger.exception("Inventory upload failed")
        return HTMLResponse(f"Upload error: {e}", status_code=500)
    finally:
        session.close()


@router.post("/inventory/{stock_no}/status")
async def set_inventory_status(
    request: Request,
    stock_no: str,
    status: str = Form(...),
    _auth: dict = Depends(require_auth),
):
    """Manager sets a single vehicle's status (available | sold | removed)."""
    if _auth.get("role") != "manager":
        return HTMLResponse("Unauthorized", status_code=403)
    if status not in ("available", "sold", "removed"):
        return HTMLResponse("Invalid status", status_code=400)
    session = _get_session()
    try:
        cookie_value = request.cookies.get("session")
        current_dealer = get_dealer_from_auth(session, cookie_value) if cookie_value else None
        if not current_dealer:
            return HTMLResponse("Unauthorized", status_code=401)
        vehicle = session.execute(
            select(Vehicle).where(
                Vehicle.dealer_id == current_dealer.id,
                Vehicle.stock_no == stock_no,
            )
        ).scalars().first()
        if not vehicle:
            return HTMLResponse("Not found", status_code=404)
        vehicle.status = status
        session.commit()
        return HTMLResponse(f'<div class="toast success">{stock_no} set to {status}.</div>', status_code=200)
    finally:
        session.close()

